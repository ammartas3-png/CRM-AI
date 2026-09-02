import csv
import io
from typing import Any

import openpyxl


def xlsx_bytes_to_csv_text(file_bytes: bytes, sheet_index: int = 0) -> str:
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    if not sheet_names:
        raise ValueError("Excel file has no sheets.")

    sheet = workbook[sheet_names[sheet_index]]
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")

    for row in sheet.iter_rows(values_only=True):
        writer.writerow(["" if cell is None else cell for cell in row])

    workbook.close()
    return output.getvalue()


def normalize_upload_to_csv(
    file_name: str,
    file_bytes: bytes,
    mime_type: str = "",
) -> tuple[str, str, bytes]:
    lowered_name = (file_name or "").lower()
    lowered_mime = (mime_type or "").lower()

    if lowered_name.endswith(".csv") or "text/csv" in lowered_mime:
        csv_text = file_bytes.decode("utf-8-sig", errors="replace")
        csv_name = file_name if lowered_name.endswith(".csv") else f"{file_name}.csv"
        return csv_name, csv_text, csv_text.encode("utf-8")

    if lowered_name.endswith((".xlsx", ".xls")) or "spreadsheet" in lowered_mime:
        csv_text = xlsx_bytes_to_csv_text(file_bytes)
        base_name = file_name.rsplit(".", 1)[0] if "." in file_name else file_name
        csv_name = f"{base_name}.csv"
        return csv_name, csv_text, csv_text.encode("utf-8")

    raise ValueError("Unsupported file type. Upload .xlsx or .csv only.")


def csv_preview_rows(csv_text: str, limit: int = 5) -> list[list[Any]]:
    reader = csv.reader(io.StringIO(csv_text))
    return [row for _, row in zip(range(limit), reader)]
